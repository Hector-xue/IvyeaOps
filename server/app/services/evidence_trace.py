"""证据溯源：从一条结论回到它依据的那几行原始数据。

统一结论契约把「指标 / 数值 / 时间窗 / 来源」写进了证据，这已经比一段散文强得多。
但用户真正想问的下一句永远是**"你从哪儿看出来的"** —— 在他决定要不要照着改
真实投放之前，他要能翻到源报表里的那几行，自己看一眼。

能做到这一点，"带证据"才不是一个说法。做不到的话，证据页上的数字和模型编的
数字在界面上长得一模一样。

设计上的两条
------------
* **原样返回，不加工**。这里回的是源文件里那几行的原文，不做单位换算、不做
  重新聚合 —— 一旦加工，用户核对的就不再是原始数据，而是我们的第二次解读。
* **匹配不到要明说**。宁可回"在源数据里没找到这一行"，也不能回一个空列表让
  界面显示成"没有证据"。前者是溯源失败，后者是结论没有依据 —— 两件完全不同
  的事，混在一起会让人对整份报告失去判断力。
"""
from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("ivyea.services.evidence_trace")

MAX_ROWS = 40          # 一条结论对应几十行就够看了，再多是在让人滚动而不是核对


def _iter_source_rows(path: Path) -> Tuple[List[str], List[List[Any]]]:
    """读一个源文件，返回 (表头, 行)。csv 与 xlsx 都认。"""
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            header = next(reader, [])
            return header, [r for r in reader if any(str(c).strip() for c in r)]

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in (next(rows, ()) or ())]
        body = [list(r) for r in rows if any(c is not None and str(c).strip() for c in r)]
        return header, body
    finally:
        wb.close()


def trace(source_paths: List[Path], target: str, *,
          limit: int = MAX_ROWS) -> Dict[str, Any]:
    """在源文件里找出与 ``target`` 有关的行。

    匹配用**不区分大小写的包含**：结论里的对象（搜索词、关键词、ASIN）在报表里
    通常就是某一列的原文，但大小写和前后空白经常对不上。
    """
    needle = (target or "").strip().lower()
    if not needle:
        return {"ok": False, "reason": "没有给出要溯源的对象", "rows": []}

    for path in source_paths:
        if not path.is_file():
            continue
        try:
            header, body = _iter_source_rows(path)
        except Exception as exc:  # noqa: BLE001 — 一个文件读不动不该毁掉整次溯源
            logger.warning("源文件读取失败 %s：%s", path.name, exc)
            continue

        hits = [r for r in body
                if any(needle in str(c).strip().strip('"').lower() for c in r)]
        if hits:
            return {
                "ok": True,
                "file": path.name,
                "columns": header,
                "rows": [[("" if c is None else c) for c in r] for r in hits[:limit]],
                "total": len(hits),
                "truncated": len(hits) > limit,
            }

    # **明说找不到**，而不是回空列表。见模块开头第二条。
    return {
        "ok": False,
        "reason": f"在源数据里没有找到包含「{target}」的行。"
                  f"可能是这条结论跨多份报表聚合而来，或源文件已被替换。",
        "rows": [],
    }


def trace_ad_audit(job_id: str, target: str) -> Optional[Dict[str, Any]]:
    """广告审计任务的溯源。任务不存在回 None（让路由回 404）。"""
    from app.services.ad_audit import _job_dir, get_job
    if not get_job(job_id):
        return None
    jd = _job_dir(job_id)
    sources = sorted(jd.glob("source_*.csv")) + sorted(jd.glob("source_*.xlsx"))
    return trace(sources, target)
